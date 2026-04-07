from openpyxl import load_workbook


def _extract_xlsx(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheets = []

    for sheet in workbook.worksheets:
        rows = []
        for row in sheet.iter_rows(values_only=True):
            cells = [str(cell).strip() for cell in row if cell not in (None, "")]
            if cells:
                rows.append("\t".join(cells))

        if rows:
            sheets.append(f"# sheet {sheet.title}\n" + "\n".join(rows))

    workbook.close()
    return "\n\n==========\n\n".join(sheets)
